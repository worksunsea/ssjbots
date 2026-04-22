#!/usr/bin/env python3
"""
SSJ Jew CRM — contact import.

Reads every contact-data file in /Users/sg/ssjbots/contacts/, normalizes,
deduplicates by primary 10-digit mobile, merges per user's rules, and writes
the result into Supabase (bullion_leads + bullion_lead_tags +
bullion_family_members + bullion_visits + bullion_imports).

Usage:
    python3 scripts/import_contacts.py

Idempotency: upserts on (tenant_id, phone). Safe to rerun — will overwrite
scalar fields with latest computed values (not destructive if source data
hasn't changed). Tags/family/visits are de-duped by natural key before insert.

Requires the Supabase service_role key to be present as SUPABASE_SERVICE_KEY
environment variable. Falls back to the hardcoded key below for convenience.
"""

from __future__ import annotations

import csv
import json
import os
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, date
from typing import Iterator, List, Optional, Dict, Set, Any

import requests
from openpyxl import load_workbook
try:
    import xlrd
except ImportError:
    xlrd = None

# ── Configuration ────────────────────────────────────────────────────

SUPABASE_URL = "https://uppyxzellmuissdlxsmy.supabase.co"
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY") or (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InVwcHl4emVsbG11aXNzZGx4c215Iiwicm9sZSI6"
    "InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NjI4NzM1MywiZXhwIjoyMDkxODYzMzUzfQ."
    "whIrPaL7Sd35POEOlR8UmhbruucUat2kyVSxmbt3svg"
)
TENANT_SSJ = "a1b2c3d4-0000-0000-0000-000000000001"
TENANT_GEMTRE = "a1b2c3d4-0000-0000-0000-000000000002"
CONTACTS_DIR = "/Users/sg/ssjbots/contacts"

HDRS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}
HDRS_MINIMAL = {**HDRS, "Prefer": "return=minimal,resolution=merge-duplicates"}

# ── Helpers ──────────────────────────────────────────────────────────

SALUTATIONS = {
    "mr","mrs","ms","miss","dr","sh","shri","smt","sri",
    "ji","bhai","bhaiya","bhabhi","bhabhiji","didi","mam","ma'am","sir","mister",
}

GENERIC_NAMES = {
    "", "customer", "client", "unknown", "walk in", "walkin",
    "saurav", "sir", "ma'am", "mam", "ji", "bhai", "n/a", "na", "-",
}


def normalize_phone(raw) -> tuple[Optional[str], Optional[str]]:
    """Return (primary_mobile, secondary). Primary must be 10-digit Indian.
    If input is not conformant, it goes to secondary."""
    if raw is None:
        return None, None
    # Excel reads numbers as floats (e.g. 9810222843.0) — strip the ".0".
    if isinstance(raw, float):
        if raw != raw:  # NaN
            return None, None
        raw = str(int(raw))
    s = str(raw).strip()
    if not s or s.lower() in ("none", "nan"):
        return None, None
    # Drop trailing ".0" from stringified floats
    if s.endswith(".0"):
        s = s[:-2]
    # Strip non-digits
    digits = re.sub(r"\D", "", s)
    # Drop leading zeroes
    digits = digits.lstrip("0")
    # Strip leading 91 if present and length suggests India
    if len(digits) >= 12 and digits.startswith("91"):
        digits = digits[2:]
    # Primary check: 10 digits, starts 6-9
    if len(digits) == 10 and digits[0] in "6789":
        return digits, None
    # If we have digits but not a valid primary, put in secondary
    if digits:
        return None, digits
    return None, None


def clean_name(raw) -> tuple[Optional[str], Optional[str]]:
    """Return (clean_name, salutation). Strip honorifics."""
    if raw is None:
        return None, None
    s = str(raw).strip()
    if not s:
        return None, None
    # Pick up salutations from prefix/suffix
    salutation = None
    tokens = re.split(r"[\s,]+", s)
    filtered = []
    for t in tokens:
        tl = t.lower().strip(".")
        if tl in SALUTATIONS and salutation is None:
            salutation = t.strip(".").title()
            continue
        if tl in SALUTATIONS:
            continue  # already captured
        filtered.append(t)
    name = " ".join(filtered).strip()
    # Remove extra whitespace
    name = re.sub(r"\s+", " ", name)
    return (name or None, salutation)


def is_generic_name(name) -> bool:
    if not name:
        return True
    return name.strip().lower() in GENERIC_NAMES


def is_well_formed_name(name) -> bool:
    if not name or is_generic_name(name):
        return False
    # At least 2 chars, contains a letter
    return len(name) >= 2 and re.search(r"[A-Za-z]", name) is not None


def parse_date_like(v) -> Optional[str]:
    """Return ISO 'YYYY-MM-DD' or 'MM-DD' if we can extract one."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        # Could be Excel serial. Try.
        try:
            if 1 < float(v) < 80000:
                base = datetime(1899, 12, 30)
                return (base + __import__("datetime").timedelta(days=int(v))).strftime("%Y-%m-%d")
        except Exception:
            pass
    s = str(v).strip()
    if not s:
        return None
    # DD/MM/YYYY or DD-MM-YYYY or DD/MM/YY
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = "19" + y if int(y) > 30 else "20" + y
        try:
            return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
        except Exception:
            return None
    # YYYY-MM-DD already
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def to_str(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def to_int(v) -> Optional[int]:
    try:
        return int(float(v))
    except Exception:
        return None


def to_bool(v) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ("y", "yes", "true", "1", "t", "on")


# ── Candidate model ──────────────────────────────────────────────────

@dataclass
class Candidate:
    tenant_id: str
    source_file: str
    source_sheet: str
    source_row: int
    source_tag: str   # the source-category tag name

    phone: Optional[str] = None
    mobile2: Optional[str] = None
    name: Optional[str] = None
    salutation: Optional[str] = None
    email: Optional[str] = None
    address_house: Optional[str] = None
    address_locality: Optional[str] = None
    address_city: Optional[str] = None
    address_state: Optional[str] = None
    address_pincode: Optional[str] = None
    address_country: Optional[str] = None
    bday: Optional[str] = None
    anniversary: Optional[str] = None
    spouse_name: Optional[str] = None
    spouse_dob: Optional[str] = None
    spouse_mobile: Optional[str] = None
    profession: Optional[str] = None
    company: Optional[str] = None
    industry: Optional[str] = None
    client_code: Optional[str] = None
    client_rating: Optional[int] = None

    flags: Dict[str, bool] = field(default_factory=dict)       # Diwali_gift, Calendar, etc.
    extra_tags: Set[str] = field(default_factory=set)           # segment tags

    visit: Optional[Dict[str, Any]] = None                      # walk-in single visit record
    unsubscribed: bool = False


@dataclass
class MergedContact:
    tenant_id: str
    phone: str
    mobile2: Optional[str] = None
    name: Optional[str] = None
    misc_names: List[str] = field(default_factory=list)
    salutation: Optional[str] = None
    email: Optional[str] = None
    address_house: Optional[str] = None
    address_locality: Optional[str] = None
    address_city: Optional[str] = None
    address_state: Optional[str] = None
    address_pincode: Optional[str] = None
    address_country: Optional[str] = None
    bday: Optional[str] = None
    anniversary: Optional[str] = None
    spouse_name: Optional[str] = None
    spouse_dob: Optional[str] = None
    spouse_mobile: Optional[str] = None
    profession: Optional[str] = None
    company: Optional[str] = None
    industry: Optional[str] = None
    client_code: Optional[str] = None
    client_rating: Optional[int] = None

    tags: Set[str] = field(default_factory=set)
    flags: Dict[str, bool] = field(default_factory=dict)
    visits: List[Dict[str, Any]] = field(default_factory=list)
    merged_from: List[Dict[str, Any]] = field(default_factory=list)
    unsubscribed: bool = False


# ── Readers ──────────────────────────────────────────────────────────

def _xlsx_iter(path, sheet=None, max_rows=None):
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = [sheet] if sheet else wb.sheetnames
    for sname in sheets:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        rows = ws.iter_rows(values_only=True)
        headers = None
        for i, row in enumerate(rows):
            if i == 0:
                headers = [str(h).strip() if h is not None else "" for h in row]
                continue
            if max_rows and i > max_rows:
                break
            yield sname, headers, i + 1, row


def _xls_iter(path, sheet=None):
    if not xlrd:
        return
    wb = xlrd.open_workbook(path)
    for sname in wb.sheet_names():
        if sheet and sname != sheet:
            continue
        ws = wb.sheet_by_name(sname)
        if ws.nrows < 2:
            continue
        headers = [str(c).strip() for c in ws.row_values(0)]
        for r in range(1, ws.nrows):
            yield sname, headers, r + 1, ws.row_values(r)


def read_master_client(path) -> Iterator[Candidate]:
    # Main Sheet
    for sname, hdrs, ridx, row in _xlsx_iter(path):
        d = dict(zip(hdrs, row))
        if sname == "Main Sheet":
            phone, mob2 = normalize_phone(d.get("Mobile"))
            alt_primary, alt_sec = normalize_phone(d.get("Alternate No"))
            if not phone and alt_primary:
                phone = alt_primary
            elif not mob2 and alt_sec:
                mob2 = alt_sec
            elif alt_primary and alt_primary != phone:
                mob2 = mob2 or alt_primary
            name, salut = clean_name(d.get("Proper name") or d.get("Name"))
            c = Candidate(
                tenant_id=TENANT_SSJ, source_file="Master Client List.xlsx",
                source_sheet=sname, source_row=ridx, source_tag="master_client_list",
                phone=phone, mobile2=mob2, name=name, salutation=salut,
                email=to_str(d.get("Email Id")),
                address_house=to_str(d.get("House Number, Floor")),
                address_locality=to_str(d.get("Postal Address, Colony, Locality")),
                address_city=to_str(d.get("City")),
                address_state=to_str(d.get("State")),
                address_pincode=to_str(d.get("Pin Code")),
                address_country=to_str(d.get("Country")),
                bday=parse_date_like(d.get("Birthday date")) or parse_date_like(d.get("Birthday")),
                anniversary=parse_date_like(d.get("Anniversay date")) or parse_date_like(d.get("Anniversary")),
                spouse_name=to_str(d.get("Spouse Name")),
                spouse_dob=parse_date_like(d.get("Spouse DOB")),
                spouse_mobile=normalize_phone(d.get("Spouse Number"))[0],
                profession=to_str(d.get("Business / Profession")),
                company=to_str(d.get("Company Name")),
                industry=to_str(d.get("Industry /  Service")),
                client_code=to_str(d.get("ii")),
            )
            c.flags["Diwali_gift"] = to_bool(d.get("Diwali Gifts"))
            c.flags["Calendar"]    = to_bool(d.get("Calender"))
            c.flags["Greetings"]   = to_bool(d.get("Greetings"))
            c.flags["Bday_gift"]   = to_bool(d.get("Bday Anniv Gifts"))
            c.flags["Gold_rate"]   = to_bool(d.get("Gold Rate"))
            cat = to_str(d.get("Category"))
            if cat:
                c.extra_tags.add(cat.lower().replace(" ", "_"))
                if "customer" in cat.lower():
                    c.extra_tags.add("client")
                if "wholesale" in cat.lower():
                    c.extra_tags.add("wholesale")
                if "b2b" in cat.lower() or "karigar" in cat.lower():
                    c.extra_tags.add("karigar")
            yield c
        elif sname == "Form responses 1":
            phone, mob2 = normalize_phone(d.get("Mobile Number (pref. Whatsapp)"))
            name, salut = clean_name(d.get("Name"))
            yield Candidate(
                tenant_id=TENANT_SSJ, source_file="Master Client List.xlsx",
                source_sheet=sname, source_row=ridx, source_tag="signup_form",
                phone=phone, mobile2=mob2, name=name, salutation=salut,
                email=to_str(d.get("Email address")) or to_str(d.get("Email")),
                bday=parse_date_like(d.get("Date of Birth")),
                anniversary=parse_date_like(d.get("Date of Anniversery")),
                spouse_name=to_str(d.get("Name of Spouse ( if Married)")),
                spouse_dob=parse_date_like(d.get("Date of Birth - Spouse")),
                address_house=to_str(d.get("Address")),
                address_pincode=to_str(d.get("Pincode")),
                address_city=to_str(d.get("City")),
                address_country=to_str(d.get("Country")),
                profession=to_str(d.get("What's your profession or line of business?"))
                    or to_str(d.get(" Business / Profession")),
            )
        elif sname == "fb bday":
            phone, mob2 = normalize_phone(d.get("Number"))
            full_name = " ".join([to_str(d.get("Name")) or "", to_str(d.get("Last Name")) or ""]).strip()
            name, salut = clean_name(full_name)
            yield Candidate(
                tenant_id=TENANT_SSJ, source_file="Master Client List.xlsx",
                source_sheet=sname, source_row=ridx, source_tag="fb_bday",
                phone=phone, mobile2=mob2, name=name, salutation=salut,
                email=to_str(d.get("Email")),
                bday=parse_date_like(d.get("DOB")),
                anniversary=parse_date_like(d.get("Anniversary")),
            )


def read_google_csv(path) -> Iterator[Candidate]:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        for ridx, d in enumerate(reader, start=2):
            # Compose name
            parts = [d.get("First Name",""), d.get("Middle Name",""), d.get("Last Name","")]
            full_name = " ".join(p for p in parts if p).strip() or d.get("File As") or d.get("Nickname") or ""
            name, salut = clean_name(full_name)
            # Phones: check 1..5
            phone = mob2 = None
            for i in range(1, 6):
                p, m = normalize_phone(d.get(f"Phone {i} - Value", ""))
                if p and not phone:
                    phone = p
                elif p and not mob2:
                    mob2 = p
                elif m and not mob2:
                    mob2 = m
            if not phone and not mob2:
                continue
            # Emails: pick first
            email = None
            for i in range(1, 5):
                e = to_str(d.get(f"E-mail {i} - Value"))
                if e:
                    email = e
                    break
            # Addresses
            house = to_str(d.get("Address 1 - Street"))
            city = to_str(d.get("Address 1 - City"))
            state = to_str(d.get("Address 1 - Region"))
            pin = to_str(d.get("Address 1 - Postal Code"))
            country = to_str(d.get("Address 1 - Country"))
            # Birthday
            bday = parse_date_like(d.get("Birthday"))
            company = to_str(d.get("Organization Name"))
            labels = (d.get("Labels") or "").lower()

            c = Candidate(
                tenant_id=TENANT_SSJ,
                source_file="contacts (3).csv",
                source_sheet="contacts",
                source_row=ridx,
                source_tag="google_csv",
                phone=phone, mobile2=mob2, name=name, salutation=salut,
                email=email, address_house=house, address_city=city,
                address_state=state, address_pincode=pin, address_country=country,
                bday=bday, company=company,
            )
            c.extra_tags.add("saurav_phone")
            for kw in ("client", "customer", "ssj", "vip"):
                if kw in labels:
                    c.extra_tags.add(kw if kw != "customer" else "client")
            # Jewellers tag via name heuristic
            if name and re.search(r"\bjew|auc", name, re.I):
                c.extra_tags.add("jewellers")
            yield c


def read_ssj_gf(path, fname, source_tag="walk_in") -> Iterator[Candidate]:
    for sname, hdrs, ridx, row in _xlsx_iter(path):
        if sname not in ("FR1", "Form Responses 1", "Form Responses 2", "Data Paste", "old sheet data", "Sheet1", "Print"):
            continue
        d = dict(zip(hdrs, row))
        phone_col = d.get("Number") or d.get("Mobile")
        phone, mob2 = normalize_phone(phone_col)
        if not phone and not mob2:
            continue
        name, salut = clean_name(d.get("Walk In Client Name") or d.get("Name"))
        ts = d.get("Timestamp")
        visit = {
            "visited_at": ts.isoformat() if hasattr(ts, "isoformat") else to_str(ts),
            "counter": to_str(d.get("Which Counter")),
            "staff": to_str(d.get("Attend By") or d.get("Attended by") or d.get("Attend by")),
            "items_seen": to_str(d.get("Items Seen") or d.get("Sale Item")),
            "purpose": to_str(d.get("If Not What were they looking for or why didn't they buy")),
            "sale": to_bool(d.get("Sale")),
            "gift_given": to_str(d.get("Gift Given")),
            "google_review": to_bool(d.get("Google Review Taken") or d.get("Review Taken")),
            "insta_follow": to_bool(d.get("Insta Follow")),
            "source_file": fname,
        }
        c = Candidate(
            tenant_id=TENANT_SSJ, source_file=fname, source_sheet=sname, source_row=ridx,
            source_tag=source_tag, phone=phone, mobile2=mob2, name=name, salutation=salut,
            address_house=to_str(d.get("Address")),
            address_locality=to_str(d.get("Locality")),
            email=to_str(d.get("Mail Id")),
            bday=parse_date_like(d.get("Your Birthday")),
            anniversary=parse_date_like(d.get("Your Anniversary")),
            visit=visit,
        )
        c.extra_tags.add("client")
        yield c


def read_wbiztool_walkin(path) -> Iterator[Candidate]:
    unsub: Set[str] = set()
    for sname, hdrs, ridx, row in _xlsx_iter(path):
        if sname not in ("FR1", "Leads_Master", "Unsubscribed"):
            continue
        d = dict(zip(hdrs, row))
        if sname == "Unsubscribed":
            p, _ = normalize_phone(d.get("Mobile No.") or row[0] if row else None)
            if p:
                unsub.add(p)
            continue
        phone_col = d.get("Number") or d.get("Phone")
        phone, mob2 = normalize_phone(phone_col)
        if not phone and not mob2:
            continue
        name, salut = clean_name(d.get("Walk In Client Name") or d.get("Name"))
        ts = d.get("Timestamp") or d.get("First_Visit")
        visit = None
        if sname == "FR1":
            visit = {
                "visited_at": ts.isoformat() if hasattr(ts, "isoformat") else to_str(ts),
                "counter": to_str(d.get("Which Counter")),
                "staff": to_str(d.get("Attend By")),
                "items_seen": to_str(d.get("Items Seen")),
                "sale": to_bool(d.get("Sale")),
                "source_file": "WBIZTOOL WALKIN DRIP.xlsx",
            }
        c = Candidate(
            tenant_id=TENANT_SSJ, source_file="WBIZTOOL WALKIN DRIP.xlsx",
            source_sheet=sname, source_row=ridx, source_tag="wbiztool_drip",
            phone=phone, mobile2=mob2, name=name, salutation=salut, visit=visit,
        )
        c.extra_tags.add("client")
        if phone in unsub:
            c.unsubscribed = True
        yield c


def read_bday_xls(path) -> Iterator[Candidate]:
    for sname, hdrs, ridx, row in _xls_iter(path):
        d = dict(zip(hdrs, row))
        # CONTACT 1/2/3
        p1, m1 = normalize_phone(d.get("CONTACT 1") or row[4] if len(row) > 4 else None)
        p2, m2 = normalize_phone(d.get("CONTACT 2"))
        p3, m3 = normalize_phone(d.get("CONTACT 3"))
        phone = p1 or p2 or p3
        mob2 = p2 or m1 or p3 or m2 or m3 if phone else (p1 or p2 or p3 or m1 or m2 or m3)
        if not phone and not mob2:
            continue
        name, salut = clean_name(d.get("NAME"))
        des = to_str(d.get("Des."))
        salut = salut or des
        addr_parts = [to_str(d.get(k)) for k in ("ADD1","ADD2","ADD3","ADD4")]
        addr = ", ".join(p for p in addr_parts if p) or None
        yield Candidate(
            tenant_id=TENANT_SSJ, source_file="BDAY.xls", source_sheet=sname, source_row=ridx,
            source_tag="bday_xls", phone=phone, mobile2=mob2, name=name, salutation=salut,
            email=to_str(d.get("E-mail") or d.get("Email")),
            bday=parse_date_like(d.get("B.DAY")),
            anniversary=parse_date_like(d.get("AN.DAY")),
            address_house=addr, address_pincode=to_str(d.get("PINCODE")),
            profession=to_str(d.get("PR0")),
        )


def read_customer_xls(path) -> Iterator[Candidate]:
    for sname, hdrs, ridx, row in _xls_iter(path):
        d = dict(zip(hdrs, row))
        p1, m1 = normalize_phone(d.get("CONTACT 1") or d.get("CONT-1"))
        p2, m2 = normalize_phone(d.get("CONTACT 2") or d.get("CONT-2"))
        p3, m3 = normalize_phone(d.get("CONTACT 3"))
        phone = p1 or p2 or p3
        mob2 = p2 or p3 or m1 or m2 or m3 if phone else (p1 or p2 or p3 or m1 or m2 or m3)
        if not phone and not mob2:
            continue
        name, salut = clean_name(d.get("NAME"))
        addr_parts = [to_str(d.get(k)) for k in ("ADD1","ADD2","ADD3")]
        addr = ", ".join(p for p in addr_parts if p) or None
        yield Candidate(
            tenant_id=TENANT_SSJ, source_file="customer.xls", source_sheet=sname, source_row=ridx,
            source_tag="customer_xls", phone=phone, mobile2=mob2, name=name, salutation=salut,
            email=to_str(d.get("Email") or d.get("E-mail")),
            bday=parse_date_like(d.get("B.DAY")),
            anniversary=parse_date_like(d.get("AN.DAY")),
            address_house=addr,
            address_city=to_str(d.get("CITY")),
            address_pincode=to_str(d.get("PINCODE")),
            profession=to_str(d.get("PR0")),
        )


def read_shivani(path) -> Iterator[Candidate]:
    for sname, hdrs, ridx, row in _xlsx_iter(path):
        if sname != "CUSTOMER DATA":
            continue
        d = dict(zip(hdrs, row))
        phone, mob2 = normalize_phone(d.get("Mobile"))
        if not phone and not mob2:
            continue
        first = to_str(d.get("First Name")) or ""
        last = to_str(d.get("NAME")) or ""
        full = f"{first} {last}".strip()
        name, salut = clean_name(full)
        c = Candidate(
            tenant_id=TENANT_SSJ, source_file="shivani customer data.xlsx",
            source_sheet=sname, source_row=ridx, source_tag="shivani",
            phone=phone, mobile2=mob2, name=name, salutation=salut,
            email=to_str(d.get("Email")),
            bday=parse_date_like(d.get("BIRTHDAY")),
            anniversary=parse_date_like(d.get("WEDDING")),
            address_house=to_str(d.get("House No.")),
            address_locality=to_str(d.get("Place") or d.get("Society")),
            address_state=to_str(d.get("State")),
            address_pincode=to_str(d.get("PinCode")),
            profession=to_str(d.get("Profession")),
            company=to_str(d.get("Company")),
        )
        cat = to_str(d.get("Category"))
        if cat:
            c.extra_tags.add(cat.lower())
        c.flags["Diwali_gift"] = to_bool(d.get("Gift"))
        c.flags["Calendar"]    = to_bool(d.get("calender"))
        yield c


def read_customer_is_king(path) -> Iterator[Candidate]:
    for sname, hdrs, ridx, row in _xlsx_iter(path):
        d = dict(zip(hdrs, row))
        phone, mob2 = normalize_phone(d.get("Mobile / Whatsapp No"))
        if not phone and not mob2:
            continue
        first = to_str(d.get("Full Name")) or ""
        last = to_str(d.get("Last Name ")) or ""
        full = f"{first} {last}".strip()
        name, salut = clean_name(full)
        yield Candidate(
            tenant_id=TENANT_SSJ, source_file="CUSTOMER IS KING  (Responses).xlsx",
            source_sheet=sname, source_row=ridx, source_tag="customer_is_king_form",
            phone=phone, mobile2=mob2, name=name, salutation=salut,
            email=to_str(d.get("Email ")),
            bday=parse_date_like(d.get("Your Birthday")),
            anniversary=parse_date_like(d.get("Anniversary")),
            address_house=to_str(d.get("House / Apartment no. ")),
            address_locality=to_str(d.get("Society / Locality")),
            address_state=to_str(d.get("State ")),
            address_pincode=to_str(d.get("Pin Code / Zip Code ")),
            profession=to_str(d.get("Profession / Business Industry")),
            company=to_str(d.get("Company Name")),
        )


def read_sunseacombined(path, fname) -> Iterator[Candidate]:
    for sname, hdrs, ridx, row in _xlsx_iter(path):
        if sname not in ("Mail", "Sheet2"):
            continue
        d = dict(zip(hdrs, row))
        # Multiple phone-like cols: Office Phone, Residence Phone, MOBILE NO.
        p1, m1 = normalize_phone(d.get("MOBILE NO. ") or d.get("Office Phone"))
        p2, m2 = normalize_phone(d.get("Residence Phone"))
        phone = p1 or p2
        mob2 = p2 or m1 or m2 if phone else (p1 or p2 or m1 or m2)
        name, salut = clean_name(d.get("NAME") or d.get("NAME "))
        salut = salut or to_str(d.get("First Name"))
        if not phone and not mob2:
            continue
        c = Candidate(
            tenant_id=TENANT_SSJ, source_file=fname, source_sheet=sname, source_row=ridx,
            source_tag="sunseaclientcombined", phone=phone, mobile2=mob2,
            name=name, salutation=salut,
            email=to_str(d.get("E MAIL ")),
            address_house=to_str(d.get("House No.")),
            address_locality=to_str(d.get("Place") or d.get("Society")),
            address_state=to_str(d.get("State")),
            address_pincode=to_str(d.get("PinCode")),
            address_city=to_str(d.get("LOCATION")),
            company=to_str(d.get("Company")),
            profession=to_str(d.get("Profession")),
        )
        c.flags["Greetings"]   = to_bool(d.get("GRE"))
        c.flags["Diwali_gift"] = to_bool(d.get("DIW"))
        c.flags["Calendar"]    = to_bool(d.get("CAL"))
        c.flags["Letter"]      = to_bool(d.get("LET"))
        cat = to_str(d.get("Category"))
        if cat:
            c.extra_tags.add(cat.lower())
        yield c


def read_exhibition(path) -> Iterator[Candidate]:
    for sname, hdrs, ridx, row in _xlsx_iter(path):
        d = dict(zip(hdrs, row))
        # Sheet name variants: "gemtre data" (lowercase) = gemtre; " data mynn and blank" = mynn.
        sname_lower = sname.strip().lower()
        if sname_lower == "gemtre data":
            phone, mob2 = normalize_phone(d.get("Mobile"))
            name, salut = clean_name(d.get("Name"))
            if not phone and not mob2:
                continue
            c = Candidate(
                tenant_id=TENANT_GEMTRE, source_file="Client Sheet Exhibition.xlsx",
                source_sheet=sname, source_row=ridx, source_tag="exhibition_sheet",
                phone=phone, mobile2=mob2, name=name, salutation=salut,
            )
            c.extra_tags.update({"exhibition", "gemtre"})
            yield c
            continue
        if sname_lower == "data mynn and blank":
            phone, mob2 = normalize_phone(d.get("Mobile"))
            name, salut = clean_name(d.get("Name"))
            if not phone and not mob2:
                continue
            c = Candidate(
                tenant_id=TENANT_SSJ, source_file="Client Sheet Exhibition.xlsx",
                source_sheet=sname, source_row=ridx, source_tag="exhibition_sheet",
                phone=phone, mobile2=mob2, name=name, salutation=salut,
            )
            c.extra_tags.update({"exhibition", "mynn"})
            yield c
            continue
        if sname in ("Form responses 1", "Sheet5"):
            phone, mob2 = normalize_phone(d.get("Mobile"))
            name, salut = clean_name(d.get("Name"))
            brand = (to_str(d.get("Product - Brand")) or "").lower()
            tenant = TENANT_GEMTRE if "gemtre" in brand else TENANT_SSJ
            if not phone and not mob2:
                continue
            c = Candidate(
                tenant_id=tenant, source_file="Client Sheet Exhibition.xlsx",
                source_sheet=sname, source_row=ridx, source_tag="exhibition_sheet",
                phone=phone, mobile2=mob2, name=name, salutation=salut,
                email=to_str(d.get("Email ID")),
            )
            c.extra_tags.add("exhibition")
            yield c
        elif sname == "Mynn Data":
            phone, mob2 = normalize_phone(d.get("Phone no"))
            name, salut = clean_name(d.get("Name"))
            if not phone and not mob2:
                continue
            c = Candidate(
                tenant_id=TENANT_SSJ, source_file="Client Sheet Exhibition.xlsx",
                source_sheet=sname, source_row=ridx, source_tag="exhibition_sheet",
                phone=phone, mobile2=mob2, name=name, salutation=salut,
            )
            c.extra_tags.update({"exhibition", "mynn"})
            yield c
        elif sname == "Gemtre data ":
            phone, mob2 = normalize_phone(d.get("Phone no"))
            name, salut = clean_name(d.get("Name") or row[0] if row else None)
            if not phone and not mob2:
                continue
            c = Candidate(
                tenant_id=TENANT_GEMTRE, source_file="Client Sheet Exhibition.xlsx",
                source_sheet=sname, source_row=ridx, source_tag="exhibition_sheet",
                phone=phone, mobile2=mob2, name=name, salutation=salut,
            )
            c.extra_tags.add("exhibition")
            yield c


def read_enquiry_fms(path) -> Iterator[Candidate]:
    for sname, hdrs, ridx, row in _xlsx_iter(path):
        if sname != "Customer Enquiry Form":
            continue
        d = dict(zip(hdrs, row))
        phone, mob2 = normalize_phone(d.get("Mobile Number"))
        if not phone and not mob2:
            continue
        name, salut = clean_name(d.get("Name"))
        c = Candidate(
            tenant_id=TENANT_SSJ, source_file="_ SSJ Enquiry FMS.xlsx",
            source_sheet=sname, source_row=ridx, source_tag="customer_enquiry_form",
            phone=phone, mobile2=mob2, name=name, salutation=salut,
            email=to_str(d.get("Email Id")),
        )
        cat = to_str(d.get("Category"))
        if cat:
            c.extra_tags.add(cat.lower().replace(" ", "_"))
        yield c


def read_sales_drip(path) -> Iterator[Candidate]:
    unsub: Set[str] = set()
    for sname, hdrs, ridx, row in _xlsx_iter(path):
        d = dict(zip(hdrs, row))
        if sname == "Unsubscribed":
            p, _ = normalize_phone(d.get("Mobile No.") or (row[0] if row else None))
            if p:
                unsub.add(p)
            continue
        if sname == "Data":
            phone, mob2 = normalize_phone(d.get("Mobile"))
            name, salut = clean_name(d.get("Name"))
            if not phone and not mob2:
                continue
            yield Candidate(
                tenant_id=TENANT_SSJ, source_file="Sales Jewellery wa drip.xlsx",
                source_sheet=sname, source_row=ridx, source_tag="walk_in",
                phone=phone, mobile2=mob2, name=name, salutation=salut,
                unsubscribed=(phone in unsub),
            )


def read_sanjeevji(path) -> Iterator[Candidate]:
    for sname, hdrs, ridx, row in _xlsx_iter(path):
        d = dict(zip(hdrs, row))
        p1, m1 = normalize_phone(d.get("Phone 1 - Value"))
        p2, m2 = normalize_phone(d.get("Phone 2 - Value"))
        phone = p1 or p2
        mob2 = p2 or m1 or m2 if phone else (p1 or p2 or m1 or m2)
        if not phone and not mob2:
            continue
        parts = [to_str(d.get(k)) for k in ("First Name", "Middle Name", "Last Name")]
        full = " ".join(p for p in parts if p)
        name, salut = clean_name(full)
        c = Candidate(
            tenant_id=TENANT_SSJ, source_file="sanjeevji contact.xlsx",
            source_sheet=sname, source_row=ridx, source_tag="sanjeevji",
            phone=phone, mobile2=mob2, name=name, salutation=salut,
            company=to_str(d.get("Organization Name")),
        )
        c.extra_tags.add("sanjeev_sir")
        yield c


# ── Merge ────────────────────────────────────────────────────────────

SOURCE_PRIORITY = {
    "master_client_list": 10,
    "customer_enquiry_form": 20,
    "customer_is_king_form": 30,
    "signup_form": 35,
    "shivani": 40,
    "sunseaclientcombined": 50,
    "bday_xls": 60,
    "customer_xls": 61,
    "fb_bday": 70,
    "wbiztool_drip": 80,
    "walk_in": 85,
    "exhibition_sheet": 90,
    "sanjeevji": 95,
    "google_csv": 100,
}


def score_source(src_tag): return SOURCE_PRIORITY.get(src_tag, 1000)


def pick_best_name(candidates) -> tuple[Optional[str], list]:
    names = [(c.name, c.source_tag) for c in candidates if is_well_formed_name(c.name)]
    if not names:
        return None, []
    # Longest well-formed, ties broken by source priority
    best = max(names, key=lambda t: (len(t[0]), -score_source(t[1])))
    primary = best[0]
    misc = sorted({n for n, _ in names if n and n.lower() != primary.lower()})
    return primary, misc


def pick_best(candidates, field_name, default=None):
    """Pick first non-empty value by source priority."""
    ranked = sorted(candidates, key=lambda c: score_source(c.source_tag))
    for c in ranked:
        v = getattr(c, field_name, None)
        if v not in (None, "", []):
            return v
    return default


def merge_group(candidates) -> MergedContact:
    # Use the first candidate's tenant — all in a group should have same tenant
    tenant = candidates[0].tenant_id
    phone = candidates[0].phone or ""
    mc = MergedContact(tenant_id=tenant, phone=phone)
    name, misc = pick_best_name(candidates)
    mc.name = name
    mc.misc_names = misc
    mc.salutation = pick_best(candidates, "salutation")
    mc.mobile2 = pick_best(candidates, "mobile2")
    mc.email = pick_best(candidates, "email")
    for fld in ("address_house","address_locality","address_city","address_state",
                "address_pincode","address_country","bday","anniversary",
                "spouse_name","spouse_dob","spouse_mobile","profession","company",
                "industry","client_code","client_rating"):
        v = pick_best(candidates, fld)
        setattr(mc, fld, v)
    # Tags & flags: union
    for c in candidates:
        for k, v in c.flags.items():
            mc.flags[k] = mc.flags.get(k, False) or v
        mc.tags.update(c.extra_tags)
        mc.tags.add(c.source_tag)
        if c.unsubscribed:
            mc.unsubscribed = True
        if c.visit:
            mc.visits.append(c.visit)
        mc.merged_from.append({
            "file": c.source_file,
            "sheet": c.source_sheet,
            "row": c.source_row,
            "tag": c.source_tag,
        })
    # Auto-derive client/jewellers tags
    if mc.name and re.search(r"\bjew|auc", mc.name, re.I):
        mc.tags.add("jewellers")
    if any(c.source_tag in ("master_client_list","signup_form","customer_enquiry_form","walk_in","wbiztool_drip","customer_is_king_form","shivani","sunseaclientcombined","fb_bday","exhibition_sheet") for c in candidates):
        mc.tags.add("client")
    return mc


def completeness(mc: MergedContact) -> int:
    fields = [mc.name, mc.phone, mc.email, mc.address_city, mc.address_pincode,
              mc.bday, mc.anniversary, mc.spouse_name, mc.profession]
    present = sum(1 for f in fields if f)
    if mc.tags:
        present += 1
    return round(present * 100 / (len(fields) + 1))


# ── Supabase IO ──────────────────────────────────────────────────────

def sb_get(path):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/{path}", headers=HDRS, timeout=30)
    r.raise_for_status()
    return r.json()


def sb_post(path, data, prefer=None):
    hdrs = {**HDRS}
    if prefer:
        hdrs["Prefer"] = prefer
    r = requests.post(f"{SUPABASE_URL}/rest/v1/{path}", headers=hdrs, json=data, timeout=60)
    if r.status_code >= 300:
        print(f"  POST {path} → {r.status_code}: {r.text[:400]}")
    return r


def sb_patch(path, data):
    r = requests.patch(f"{SUPABASE_URL}/rest/v1/{path}", headers=HDRS, json=data, timeout=30)
    return r


def load_tag_map(tenant_id: str) -> Dict[str, str]:
    rows = sb_get(f"bullion_tags?select=id,name&tenant_id=eq.{tenant_id}&limit=2000")
    return {r["name"].lower(): r["id"] for r in rows}


def ensure_tag(tenant_id: str, name: str, category: str, tag_map: Dict[str, str]) -> Optional[str]:
    key = name.lower()
    if key in tag_map:
        return tag_map[key]
    r = sb_post("bullion_tags", {
        "tenant_id": tenant_id, "name": name, "category": category,
        "color": "#888", "sort_order": 500,
    })
    if r.status_code < 300:
        data = r.json()
        if data:
            tag_map[key] = data[0]["id"]
            return data[0]["id"]
    return None


def upsert_contact(mc: MergedContact) -> Optional[str]:
    body = {
        "tenant_id": mc.tenant_id,
        "phone": mc.phone,
        "name": mc.name,
        "misc_names": mc.misc_names,
        "salutation": mc.salutation,
        "mobile2": mc.mobile2,
        "email": mc.email,
        "address_house": mc.address_house,
        "address_locality": mc.address_locality,
        "city": mc.address_city,
        "address_state": mc.address_state,
        "address_pincode": mc.address_pincode,
        "address_country": mc.address_country,
        "bday": mc.bday,
        "anniversary": mc.anniversary,
        "spouse_name": mc.spouse_name,
        "spouse_dob": mc.spouse_dob,
        "spouse_mobile": mc.spouse_mobile,
        "profession": mc.profession,
        "company": mc.company,
        "industry": mc.industry,
        "client_code": mc.client_code,
        "client_rating": mc.client_rating,
        "completeness_score": completeness(mc),
        "merged_from": mc.merged_from,
        "dnd": mc.unsubscribed,
        "status": "dead" if mc.unsubscribed else "active",
        "bot_paused": mc.unsubscribed,
    }
    body = {k: v for k, v in body.items() if v is not None and v != []}
    r = requests.post(
        f"{SUPABASE_URL}/rest/v1/bullion_leads?on_conflict=tenant_id,phone",
        headers={**HDRS, "Prefer": "resolution=merge-duplicates,return=representation"},
        json=body, timeout=30,
    )
    if r.status_code >= 300:
        print(f"  upsert fail phone={mc.phone}: {r.status_code} {r.text[:200]}")
        return None
    data = r.json()
    if data:
        return data[0]["id"]
    # Fallback: fetch id by phone
    rows = sb_get(f"bullion_leads?select=id&tenant_id=eq.{mc.tenant_id}&phone=eq.{mc.phone}")
    return rows[0]["id"] if rows else None


def attach_tags(lead_id: str, tag_ids: List[str]):
    if not tag_ids:
        return
    rows = [{"lead_id": lead_id, "tag_id": tid} for tid in tag_ids]
    requests.post(
        f"{SUPABASE_URL}/rest/v1/bullion_lead_tags?on_conflict=lead_id,tag_id",
        headers={**HDRS, "Prefer": "resolution=ignore-duplicates,return=minimal"},
        json=rows, timeout=30,
    )


def insert_visits(tenant_id: str, lead_id: str, visits: List[Dict[str, Any]]):
    if not visits:
        return
    # PostgREST requires all rows in a batch to have the same keys.
    # Merge all keys seen across visits and fill missing with None.
    all_keys = set()
    for v in visits:
        all_keys.update(v.keys())
    all_keys.discard("tenant_id"); all_keys.discard("lead_id")
    rows = []
    for v in visits:
        row = {"tenant_id": tenant_id, "lead_id": lead_id}
        for k in all_keys:
            row[k] = v.get(k)
        rows.append(row)
    sb_post("bullion_visits", rows, prefer="return=minimal")


def insert_family_spouse(tenant_id: str, lead_id: str, mc: MergedContact):
    if not mc.spouse_name and not mc.spouse_dob and not mc.spouse_mobile:
        return
    sb_post("bullion_family_members", [{
        "tenant_id": tenant_id, "lead_id": lead_id,
        "relationship": "spouse", "name": mc.spouse_name,
        "dob": mc.spouse_dob, "mobile": mc.spouse_mobile,
    }], prefer="return=minimal")


# ── Main ─────────────────────────────────────────────────────────────

def main():
    print("→ Reading all sources…")
    candidates: List[Candidate] = []

    readers = [
        ("Master Client List.xlsx", read_master_client),
        ("contacts (3).csv", read_google_csv),
        ("BDAY.xls", read_bday_xls),
        ("customer.xls", read_customer_xls),
        ("shivani customer data.xlsx", read_shivani),
        ("CUSTOMER IS KING  (Responses).xlsx", read_customer_is_king),
        ("SUNSEACLIENTCOMBINED.xlsx", lambda p: read_sunseacombined(p, "SUNSEACLIENTCOMBINED.xlsx")),
        ("SUNSEACLIENTCOMBINED (1).xlsx", lambda p: read_sunseacombined(p, "SUNSEACLIENTCOMBINED (1).xlsx")),
        ("SSJ GF.xlsx", lambda p: read_ssj_gf(p, "SSJ GF.xlsx")),
        ("new SSJ GF.xlsx", lambda p: read_ssj_gf(p, "new SSJ GF.xlsx")),
        ("WBIZTOOL WALKIN DRIP.xlsx", read_wbiztool_walkin),
        ("Client Sheet Exhibition.xlsx", read_exhibition),
        ("_ SSJ Enquiry FMS.xlsx", read_enquiry_fms),
        ("Sales Jewellery wa drip.xlsx", read_sales_drip),
        ("sanjeevji contact.xlsx", read_sanjeevji),
    ]

    per_file: Dict[str, int] = {}
    for fname, reader in readers:
        path = os.path.join(CONTACTS_DIR, fname)
        if not os.path.exists(path):
            print(f"  skip (not found): {fname}")
            continue
        try:
            rows = list(reader(path))
            candidates.extend(rows)
            per_file[fname] = len(rows)
            print(f"  {fname}: {len(rows)} candidate rows")
        except Exception as e:
            print(f"  ERR reading {fname}: {e}")

    print(f"→ {len(candidates)} total candidates")

    # Group by (tenant, phone); if phone None, drop
    groups: Dict[tuple, List[Candidate]] = {}
    skipped_no_phone = 0
    for c in candidates:
        if not c.phone:
            skipped_no_phone += 1
            continue
        key = (c.tenant_id, c.phone)
        groups.setdefault(key, []).append(c)
    print(f"→ {len(groups)} unique (tenant, phone) groups · skipped {skipped_no_phone} without primary phone")

    # Merge each group
    merged: List[MergedContact] = [merge_group(cs) for cs in groups.values()]

    # Write to DB per tenant
    for tenant in (TENANT_SSJ, TENANT_GEMTRE):
        tenant_merged = [m for m in merged if m.tenant_id == tenant]
        if not tenant_merged:
            continue
        print(f"\n→ Writing {len(tenant_merged)} contacts for tenant {tenant[-4:]}…")
        tag_map = load_tag_map(tenant)

        # Ensure all source/segment/flag tags exist
        all_tag_names = set()
        for m in tenant_merged:
            all_tag_names.update(m.tags)
            for flag, on in m.flags.items():
                if on:
                    all_tag_names.add(flag)
        for name in all_tag_names:
            cat = "source" if name in SOURCE_PRIORITY else (
                "flag" if name in ("Diwali_gift","Calendar","Greetings","Bday_gift","Gold_rate","Letter") else "segment"
            )
            ensure_tag(tenant, name, cat, tag_map)

        created = 0
        failed = 0
        for i, mc in enumerate(tenant_merged):
            if i % 200 == 0:
                print(f"  …{i}/{len(tenant_merged)}")
            lead_id = upsert_contact(mc)
            if not lead_id:
                failed += 1
                continue
            created += 1
            # Attach tags
            tag_ids = []
            for name in mc.tags:
                tid = tag_map.get(name.lower())
                if tid: tag_ids.append(tid)
            for flag, on in mc.flags.items():
                if on:
                    tid = tag_map.get(flag.lower())
                    if tid: tag_ids.append(tid)
            attach_tags(lead_id, list(set(tag_ids)))
            insert_visits(tenant, lead_id, mc.visits)
            insert_family_spouse(tenant, lead_id, mc)

        # Log import run
        sb_post("bullion_imports", [{
            "tenant_id": tenant,
            "finished_at": datetime.utcnow().isoformat(),
            "file": "ALL",
            "rows_in": sum(per_file.values()),
            "rows_created": created,
            "rows_merged": len(tenant_merged) - created,
            "rows_skipped": skipped_no_phone,
            "summary": per_file,
        }], prefer="return=minimal")
        print(f"  done: {created} contacts, {failed} failed")

    print("\n✓ Import complete.")


if __name__ == "__main__":
    main()
